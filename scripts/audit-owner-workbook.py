"""Match Newlagoonsandnoya.xlsx owner records to canonical application villa keys.

This script is deliberately read-only. It emits a reviewed JSON manifest that
separates exact matches from duplicate or unmatched rows; a separate database
step may import only the exact records after the manifest is reviewed.
"""

import json
import re
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = Path("/home/ubuntu/upload/Newlagoonsandnoya.xlsx")
LAGOONS = ROOT / "server/data/lagoons.json"
LAGOONS_COORDS = ROOT / "client/src/data/lagoonsCoordinates.ts"
OTHER = ROOT / "server/data/aldar_other.json"
OUTPUT = ROOT / "server/data/owner_workbook_2026_08_26_audit.json"


def norm(value: Any) -> str:
    text = str(value or "").upper()
    text = text.replace("WIIDS", "WILDS").replace("\\.'", "W")
    text = text.replace("AI", "AL")
    return re.sub(r"[^A-Z0-9]", "", text)


def phone(value: Any) -> str | None:
    if value is None:
        return None
    raw = str(value).strip()
    if not raw or raw.upper() in {"NO NUMBER", "N/A", "NONE", "-"}:
        return None
    digits = re.sub(r"\D", "", raw)
    return digits or None


def walk_lagoons(value: Any):
    if isinstance(value, dict):
        if "unit_number" in value:
            yield value
        for nested in value.values():
            yield from walk_lagoons(nested)
    elif isinstance(value, list):
        for nested in value:
            yield from walk_lagoons(nested)


def sheet_rows(sheet):
    """Yield (source_key, owner, phone) across header and headerless sheets."""
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return
    first = [str(cell or "").strip().lower() for cell in values[0]]
    has_header = any("unit" in cell for cell in first)
    start = 1 if has_header else 0
    key_idx = next((idx for idx, cell in enumerate(first) if "unit" in cell), 0)
    owner_idx = next((idx for idx, cell in enumerate(first) if "owner" in cell), 1)
    phone_idx = next((idx for idx, cell in enumerate(first) if "contact" in cell or "mobile" in cell), 2)
    for row_num, row in enumerate(values[start:], start=start + 1):
        key = row[key_idx] if len(row) > key_idx else None
        owner = row[owner_idx] if len(row) > owner_idx else None
        mobile = row[phone_idx] if len(row) > phone_idx else None
        if not key or not owner:
            continue
        yield row_num, str(key).strip(), str(owner).strip(), phone(mobile)


def main() -> None:
    lagoons_source = json.loads(LAGOONS.read_text())
    coords_text = LAGOONS_COORDS.read_text()
    coords_start = coords_text.index("[", coords_text.index("="))
    coords_end = coords_text.rfind("];" ) + 1
    lagoons_coords = json.loads(coords_text[coords_start:coords_end])
    lagoons_index: dict[str, list[dict[str, str]]] = defaultdict(list)
    lagoons_phase_index: dict[tuple[str, str, str], list[dict[str, str]]] = defaultdict(list)
    for unit in lagoons_coords:
        unit_name = str(unit["unit_name"])
        parsed = re.match(r"^(AlGhaf|AlSidr|Ethir)-(\d+)-\d+$", unit_name)
        phase = re.match(r"^SL(\d+)$", str(unit.get("sl_phase") or ""))
        if not parsed or not phase:
            continue
        cluster, villa = parsed.groups()
        cluster_code = {"AlGhaf": "ALGHAF", "AlSidr": "ALSIDR", "Ethir": "ETHIR"}[cluster]
        entry = {"villa_key": f"lagoons/{unit_name}", "community": "lagoons", "source_unit": unit_name}
        lagoons_phase_index[(cluster_code, villa.zfill(3), phase.group(1).zfill(2))].append(entry)

    other = json.loads(OTHER.read_text())
    noya_index: dict[str, list[dict[str, str]]] = defaultdict(list)
    for project in other.get("projects", []):
        name = str(project.get("name", "")).lower()
        if "noya" not in name:
            continue
        for building in project.get("buildings", []):
            for unit in building.get("units", []):
                raw = str(unit.get("unit_name", ""))
                if not raw:
                    continue
                key = f"aldar-other/{project['slug']}/{building['slug']}/{raw}"
                entry = {"villa_key": key, "community": "aldar-other", "source_unit": raw, "project": project.get("name", "")}
                noya_index[norm(raw)].append(entry)
                # The workbook's Noya Luma tab omits the project prefix.
                for suffix in ("N3V", "N3TH"):
                    at = norm(raw).find(suffix)
                    if at >= 0:
                        noya_index[norm(raw)[at:]].append(entry)

    exact, ambiguous, unmatched = [], [], []
    workbook = load_workbook(WORKBOOK, data_only=True, read_only=True)
    for sheet in workbook.worksheets:
        is_lagoons = sheet.title.strip().lower() in {"al ghaf", "ethir", "al sidr"}
        index = lagoons_index if is_lagoons else noya_index
        for row_num, source_key, owner_name, owner_phone in sheet_rows(sheet):
            candidates = index.get(norm(source_key), [])
            if is_lagoons and not candidates:
                compact = norm(source_key)
                source_match = re.search(r"LAGOONS(WILDS|ALGHAF|ALSIDR|ETHIR)SL(\d+)V(\d+)$", compact)
                if source_match:
                    source_cluster = {"WILDS": "ALGHAF", "ALGHAF": "ALGHAF", "ALSIDR": "ALSIDR", "ETHIR": "ETHIR"}[source_match.group(1)]
                    candidates = lagoons_phase_index.get((source_cluster, source_match.group(3).zfill(3), source_match.group(2).zfill(2)), [])
            payload = {
                "sheet": sheet.title,
                "row": row_num,
                "source_unit": source_key,
                "owner_name": owner_name,
                "owner_phone": owner_phone,
                "candidates": candidates,
            }
            distinct = {candidate["villa_key"] for candidate in candidates}
            if len(distinct) == 1:
                candidate = candidates[0]
                exact.append({
                    **payload,
                    "villa_key": candidate["villa_key"],
                    "community": candidate["community"],
                    "project": candidate.get("project"),
                })
            elif candidates:
                ambiguous.append(payload)
            else:
                unmatched.append(payload)

    # A canonical key must not receive two different owner values from this workbook.
    conflicts = []
    by_villa: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in exact:
        by_villa[record["villa_key"]].append(record)
    approved = []
    for villa_key, records in by_villa.items():
        owners = {(r["owner_name"].strip().upper(), r.get("owner_phone") or "") for r in records}
        if len(owners) == 1:
            approved.append(records[0])
        else:
            conflicts.append({"villa_key": villa_key, "records": records, "reason": "conflicting_owner_values"})

    output = {
        "source_file": WORKBOOK.name,
        "generated_at": date.today().isoformat(),
        "approved_records": approved,
        "ambiguous_records": ambiguous,
        "unmatched_records": unmatched,
        "conflicting_records": conflicts,
        "summary": {
            "approved": len(approved),
            "lagoons": sum(1 for r in approved if r["community"] == "lagoons"),
            "noya": sum(1 for r in approved if r["community"] == "aldar-other"),
            "ambiguous": len(ambiguous),
            "unmatched": len(unmatched),
            "conflicts": len(conflicts),
        },
    }
    OUTPUT.write_text(json.dumps(output, indent=2))
    print(json.dumps(output["summary"], indent=2))


if __name__ == "__main__":
    main()
