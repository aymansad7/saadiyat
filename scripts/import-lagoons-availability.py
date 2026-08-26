from __future__ import annotations

import json
import re
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = Path("/home/ubuntu/upload/Saadiyat_Lagoons_Availability.xlsx")
LAGOONS = ROOT / "server/data/lagoons.json"
OUTPUT = ROOT / "server/data/lagoons_availability_2026_08_26.json"


def walk(value: Any):
    if isinstance(value, dict):
        if "unit_number" in value:
            yield value
        for nested in value.values():
            yield from walk(nested)
    elif isinstance(value, list):
        for nested in value:
            yield from walk(nested)


def number(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).replace("AED", "").replace(",", "").strip()
    try:
        return float(text)
    except ValueError:
        return None


def normalize_source_unit(value: str) -> tuple[str, str] | None:
    compact = re.sub(r"\s+", "", value.replace("WiIds", "Wilds"))
    match = re.search(r"(?:Lagoons-)?(.+?)-SL\d+-V-(\d+)$", compact, re.I)
    if not match:
        match = re.search(r"^SL\d+-V-(\d+)$", compact, re.I)
        return ("", match.group(1)) if match else None
    cluster = match.group(1).replace("-", "")
    cluster = {"Wilds": "AlGhaf", "AlSidr": "AlSidr", "AlGhaf": "AlGhaf"}.get(cluster, cluster)
    return cluster, match.group(2).zfill(3)


def candidate_metrics(unit: dict[str, Any]) -> tuple[float | None, int | None]:
    area = None
    for key in ("saleable_area_sqft", "built_up_sqft", "built_up_area_sqft", "total_built_up_area_sqft"):
        if unit.get(key) is not None:
            area = number(unit[key])
            break
    bedrooms = unit.get("bedrooms") or unit.get("bedroom_count") or unit.get("number_of_bedrooms")
    try:
        bedrooms = int(bedrooms) if bedrooms is not None else None
    except (TypeError, ValueError):
        bedrooms = None
    return area, bedrooms


def main() -> None:
    source = json.loads(LAGOONS.read_text())
    units = list(walk(source))
    workbook = load_workbook(WORKBOOK, data_only=True, read_only=True)
    sheet = workbook["Availability"]
    rows = list(sheet.iter_rows(values_only=True))
    header_index = next(index for index, row in enumerate(rows) if row and row[0] == "ID")
    headers = list(rows[header_index])

    listings: list[dict[str, Any]] = []
    unmatched: list[dict[str, Any]] = []
    ambiguous: list[dict[str, Any]] = []
    for raw in rows[header_index + 1 :]:
        if not raw or raw[0] is None:
            continue
        row = dict(zip(headers, raw))
        if str(row.get("Property Type", "")).lower() != "villa":
            unmatched.append({"source_id": row.get("ID"), "reason": "not_a_villa", "unit_number": row.get("Unit number")})
            continue
        source_unit = str(row.get("Unit number") or "").strip()
        normalized = normalize_source_unit(source_unit)
        if not normalized:
            unmatched.append({"source_id": row.get("ID"), "reason": "unparseable_unit", "unit_number": source_unit})
            continue
        cluster, villa = normalized
        candidates = [unit for unit in units if re.search(rf"-V-{re.escape(villa)}-\d+$", str(unit.get("unit_number", "")))]
        if cluster:
            candidates = [unit for unit in candidates if f"Lagoons-{cluster}-" in str(unit.get("unit_number", ""))]
        source_area = number(row.get("Area SQFT"))
        source_bedrooms = int(row["Number of Bedrooms"]) if row.get("Number of Bedrooms") else None
        narrowed = candidates
        if source_area is not None:
            area_matches = [unit for unit in narrowed if (metric := candidate_metrics(unit)[0]) is not None and abs(metric - source_area) <= 10]
            if area_matches:
                narrowed = area_matches
        if source_bedrooms is not None:
            bedroom_matches = [unit for unit in narrowed if candidate_metrics(unit)[1] == source_bedrooms]
            if bedroom_matches:
                narrowed = bedroom_matches

        listing = {
            "source_id": int(row["ID"]),
            "source_date": "2026-08-26",
            "source_label": "Shared availability sheet",
            "unit_number": source_unit,
            "asking_price_aed": int(number(row.get("Offering Price")) or 0),
            "status": str(row.get("Stage") or "Available"),
            "bedrooms": source_bedrooms,
            "area_sqft": source_area,
            "agent": row.get("Responsible Listing Agent"),
            "matched_villa_keys": [unit["unit_number"] for unit in narrowed],
        }
        if len(narrowed) == 1:
            listing["match_status"] = "exact"
            listings.append(listing)
        elif narrowed:
            listing["match_status"] = "ambiguous"
            ambiguous.append(listing)
        else:
            unmatched.append({"source_id": row.get("ID"), "reason": "no_dataset_candidate", "unit_number": source_unit})

    output = {
        "source_file": WORKBOOK.name,
        "source_date": "2026-08-26",
        "generated_at": date.today().isoformat(),
        "exact_listings": listings,
        "ambiguous_listings": ambiguous,
        "unmatched_rows": unmatched,
        "summary": {
            "source_rows": len(rows) - header_index - 1,
            "exact_villas": len(listings),
            "ambiguous_villas": len(ambiguous),
            "unmatched_rows": len(unmatched),
        },
    }
    OUTPUT.write_text(json.dumps(output, indent=2))
    print(json.dumps(output["summary"], indent=2))


if __name__ == "__main__":
    main()
