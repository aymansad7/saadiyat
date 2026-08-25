#!/usr/bin/env python3
"""Extract Saadiyat Reserve sale inventory from the supplied Valancia workbook."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = Path("/home/ubuntu/upload/CopyofValanciaInventory2026updating(1).xlsx")
DEFAULT_OUTPUT = ROOT / "scripts/source-data/saadiyat-reserve-inventory.json"
SOURCE_RECEIVED_AT = "2026-08-25"


def numeric(value):
    if value in (None, "", "-"):
        return None
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else round(float(value), 2)
    cleaned = str(value).replace(",", "").strip()
    try:
        number = float(cleaned)
    except ValueError:
        return None
    return int(number) if number.is_integer() else round(number, 2)


def parse_price(value) -> tuple[str | None, int | None]:
    if value is None:
        return None, None
    text = str(value).strip()
    if not text:
        return None, None
    if text.lower() == "sold":
        return "sold", None
    digits = re.sub(r"[^0-9.]", "", text)
    if not digits:
        return None, None
    return "available_for_sale", round(float(digits))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook = load_workbook(args.workbook, data_only=True)
    sheet = workbook["Buy"]
    headers = [cell.value for cell in sheet[1]]
    records = []
    current_project = None

    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = dict(zip(headers, values))
        project_value = str(row.get("Project") or "").strip()
        if project_value:
            current_project = project_value
        is_reserve = project_value == "Saadiyat Reserve" or (
            not project_value and current_project == "Saadiyat Reserve" and row.get("Unit Number") not in (None, "")
        )
        if not is_reserve:
            continue

        unit_number = numeric(row.get("Unit Number"))
        if unit_number is None:
            continue
        status, price = parse_price(row.get("Price"))
        unit_type = str(row.get("Unit Type") or "").strip()
        records.append(
            {
                "sourceSheet": "Buy",
                "sourceRow": row_number,
                "sourceReceivedAt": SOURCE_RECEIVED_AT,
                "project": "Saadiyat Reserve",
                "sourcePhase": str(row.get("Phase") or "").strip() or None,
                "unitType": unit_type,
                "unitNumber": int(unit_number),
                "landAreaSqm": numeric(row.get("Land Area (sqm)")),
                "builtUpAreaSqm": numeric(row.get("Built-Up Area")),
                "floors": str(row.get("Floors") or "").strip() or None,
                "bedrooms": numeric(row.get("Bedrooms")),
                "bathrooms": numeric(row.get("Bathrooms")),
                "parking": numeric(row.get("Parking")),
                "availability": status,
                "askingPriceAed": price,
                "negotiable": str(row.get("Negotiable") or "").strip().upper() == "YES",
                "referenceNumber": str(row.get("Reference Number") or "").strip() or None,
                "details": str(row.get("Details") or "").strip() or None,
            }
        )

    if len(records) != 9:
        raise RuntimeError(f"Expected 9 Saadiyat Reserve rows, found {len(records)}")
    if sum(record["availability"] == "available_for_sale" for record in records) != 7:
        raise RuntimeError("Expected 7 available Saadiyat Reserve rows")
    if sum(record["availability"] == "sold" for record in records) != 2:
        raise RuntimeError("Expected 2 sold Saadiyat Reserve rows")
    if sum(record["unitType"] == "Standalone Villa" for record in records) != 3:
        raise RuntimeError("Expected 3 built Phase 2 villas")

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(records, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Wrote {len(records)} Reserve rows → {args.output}")


if __name__ == "__main__":
    main()
